import openpyxl
import pandas as pd
import re

# Extract all reqired from content book creation and store it as object for using in the automation
class ContentBook:
    def __init__(self, filepath):
        self.filepath = filepath
        self.contentbook = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        self._hotel_content()
        self.close_workbook()
        
    def _hotel_content(self):
        self._address_setup_data()
        self._rooms_data()
        self._rooms_description_data()
        self._web_description_data()
        self._main_services_data()
        self._products_data()
        self._sport_leisure_data()
        self._restaurant_data()
        self._bar_data()
        self._meeting_room_data()
        self._mean_of_access_data()
        self._main_attractions_data()
        self._surrounding_attracition_data()
        self._product_lib_df()
        self._resaurant_description()
        self._all_web_adress()
                
    def close_workbook(self):
        if self.contentbook:
            self.contentbook.close()

    # Read data from workbook return list     
    def get_values(self, sheet_name=str, cell_addresses=list):
        try:
            # Select the specific sheet
            sheet = self.contentbook[sheet_name]
            cell_values =[]
            for cell in cell_addresses:
                cell_values.append(sheet[cell].value)
            return cell_values

        except Exception as e:
            print(f"An error occurred while getting cell values: {str(e)}")
            return None
        
    # Extract Capital Letter   
    def extract_capital_letters(self, input_string):
        # Define the regular expression pattern to match capital letters
        pattern = r'[A-Z]'
        matches = re.findall(pattern, input_string)
        capital_letters = ''.join(matches)
        return capital_letters
            
    # List data for address and setup page
    def _address_setup_data(self):
        # cell_addresses = ['C4', 'C6', 'J6', 'D32', 'K10', 'C34', 'J34', 'C37', 'J39', 'K43', 'K45',
        #                 'C47', 'K37', 'C41', 'C45', 'C49', 'C39', 'D41', 'D43',
        #                 'C57', 'J57', 'D12', 'J63', 'C61', 'J59', 'C59', 'K65', 'C63', 'C65']
        
        sheet_name = 'Address&Setup'
        sheet = self.contentbook[sheet_name]
        
        # data for Address and Setup page
        self.hotel_rid = sheet['K4'].value
        self.hotel_name = sheet['C4'].value
        self.brand = sheet['C6'].value
        self.chain = sheet['J6'].value
        self.hotel_commercial_name = sheet['D32'].value
        self.open_date = sheet['K10'].value
        self.address1 = sheet['C34'].value
        self.address2 = sheet['J34'].value
        self.address3 = sheet['C37'].value
        self.state = sheet['J39'].value
        self.resa_phone = sheet['K43'].value
        self.resa_fax = sheet['K45'].value
        self.logging_type = sheet['C47'].value
        self.zip_code = sheet['K37'].value
        self.phone_country_code = sheet['C41'].value
        self.hotel_email = sheet['C45'].value
        self.code_place = sheet['C49'].value
        self.city = sheet['C39'].value
        self.phone = sheet['D41'].value
        self.fax = sheet['D43'].value
        self.country = sheet['J41'].value
    
        # List data for general page
        self.construction_date = sheet['C57'].value
        self.reno_date = sheet['J57'].value
        self.distribute_tars_date = sheet['D12'].value
        self.lodging_type = sheet['J63'].value
        self.environment = sheet['C61'].value
        self.location = sheet['J59'].value
        self.currency = sheet['C59'].value
        self.nb_lifts = sheet['K65'].value
        self.nb_rooms = sheet['C63'].value
        self.nb_floors = sheet['C65'].value
        
        # rating data
        self.local_rating = sheet['C79'].value
        self.north_star = sheet['J79'].value
        
        # GM name
        gm_title = sheet['E53'].value
        gm_name = sheet['J55'].value
        gm_surname = sheet['J53'].value

        # GM name
        self.gm = self.accor_name(gm_title, gm_name, gm_surname)
        
        # GPS not send
        self.gps = sheet['K24'].value
        
        # Hotel Currency 3 digits
        self.currency_code = self.extract_currency(self.currency)
        
    def _all_web_adress(self):
        self.hotel_url = f'all.accor.com/{self.hotel_rid}'
        
    def get_chain_code(self):
        from dictionary import chain_dict
        chain_code = chain_dict.get(self.chain)
        return chain_code
    
    def get_country_code(self):
        from dictionary import country_dict
        code = country_dict.get(self.country)
        return code
    
    def remove_special_char(self, input_string):
    # Use regex to remove all non-alphanumeric characters
        result_string = re.sub(r'[^a-zA-Z0-9\s]', '', input_string)
        return result_string
    
    def extract_currency(self, input):
        pattern = r'-(.+)'
        match = re.search(pattern, input)
        if match:
            result = match.group(1).strip()
        else:
            result = None   
        return result

    def _rooms_data(self):
        df = pd.read_excel(
            self.filepath, 
            sheet_name="Roomtypes", 
            header=8, 
            dtype=str
            )
        
        df = df.dropna(subset=['TARS product code'])
        df = df.dropna(axis=1, how='all')
        df = df.drop(columns=['Unnamed: 1'])
        df = df.reset_index(drop=True)
        df = df.drop(index=0)
        columns_to_drop = [col for col in df.columns if 'Unnamed' in col]
        df = df.drop(columns=columns_to_drop)
        # Print the DataFrame using repr to show all hidden charactors
        # for column_name in df.columns:
        #    print(repr(column_name))
        self.rooms_df = df
        return df
    
    # Clean Dataframe return col code tar name and tars descritption
    def _rooms_description_data(self):
        df = self.rooms_df
        keep_cols = ['TARS product code', 
                    'Room Type Name in TARS reference language (marketing label)\nMax 94 characters', 
                    'Room Type Description in TARS reference language\nMax 250 characters']
        
        new_cols_name = {
            'TARS product code': 'room_code', 
            'Room Type Name in TARS reference language (marketing label)\nMax 94 characters': 'marketing_label', 
            'Room Type Description in TARS reference language\nMax 250 characters': 'tar_ref'
        }
        df = df.loc[:, keep_cols]
        df = df.rename(columns=new_cols_name)
        df = df.dropna()
        df = df.reset_index(drop=True)
        self.room_description_df = df
        
        return df
    
    def _web_description_data(self):
        sheet_name = 'Main services'
        sheet_data = self.contentbook[sheet_name]
        # Initialize row and column indices
        current_row = 40
        current_col = 3  # Refer to Column C
        data_list = []
        
        while current_row <= 125:
            try:
                l_value = sheet_data.cell(row=current_row, column=12).value  # Column L
                if l_value == "Yes":
                    code = str(sheet_data.cell(row=current_row, column=current_col).value)[0:6].strip()
                    description = None
                    marketing = None

                    # Check the next row for Description or Marketing
                    for i in range(1, 3):
                        cell_value = str(sheet_data.cell(row=current_row + i, column=current_col).value).split()

                        if "Description." in cell_value:
                            description = sheet_data.cell(row=current_row + i, column=4).value
                        elif "Marketing" in cell_value:
                            marketing = sheet_data.cell(row=current_row + i, column=4).value

                    # Append the data to the list of dictionaries
                    data_list.append({'Code': code, 'Description': description, 'Marketing': marketing})
                
                # go to next row
                current_row += 1
                
            # if index out of range break
            except IndexError:
                break
        # Convert the list of dictionaries to a DataFrame
        df = pd.DataFrame(data_list)
        self.web_description_df = df
        return df
    
    def _main_services_data(self):
        sheet_name = 'Main services'
        row_start = 11
        row_end = 37
        col_code = []
        col_available = []
        col_amount = []
        for i in range(row_start, row_end + 1):
            col_code.append(f'C{i}')
            col_available.append(f'J{i}')
            col_amount.append(f'K{i}')
        code = self.get_values(sheet_name, cell_addresses=col_code)
        available = self.get_values( sheet_name, cell_addresses=col_available)
        amount = self.get_values(sheet_name, cell_addresses=col_amount)
        df = pd.DataFrame({'code': code, 'available': available, 'amount': amount})
        # Clean Data frame filter only 'Yes'
        df.dropna(subset=['available'], inplace=True)
        df = df.loc[df['available'] == 'Yes']
        self.main_service_df = df
        return df
    
    def _products_data(self):
        products_df = pd.read_excel(
            self.filepath, 
            sheet_name="Other Services", 
            skiprows=9
            )
        
        # Clean data
        products_df = products_df.drop(columns=['Unnamed: 0', 'Family', 'Hotel services', 'Unnamed: 4', 'Unnamed: 5', 'Unnamed: 6',
                                'Displayed on AccorHotels.com', 'Amount', 'Unnamed: 11'])
        products_df = products_df.dropna()
        products_df = products_df[products_df['Product is present\nYes/No'] != 'No']
        products_df = products_df.rename(columns={
            'Product is present\nYes/No': 'available',
            'Paying\nYes/No': 'paying'
            })
        self.products_df = products_df
        return products_df
    
    def _sport_leisure_data(self):
        sheet = self.contentbook['Sports&Leisure']
        sports_leisure_data = []
        
        # Skip empty Cell
        row_index = 13
        i = 0
        while True:
            product_available = str(sheet[f'H{row_index + i}'].value) 
            product_code = str(sheet[f'C{row_index + i}'].value).strip()  
            on_site = str(sheet[f'I{row_index + i}'].value) 
            paying = str(sheet[f'J{row_index + i}'].value) 
            
            if product_available == 'Yes':
                data_entry = {
                    'product_code': product_code,
                    'on_site': on_site,
                    'paying': paying
                    }
                    
                sports_leisure_data.append(data_entry)
            i += 1
                
            if product_code == 'None':
                break
                
        df = pd.DataFrame(sports_leisure_data)
        self.sport_leisure_df = df
        return df
    
    def count_rows(self, sheet, column_letter, start_row):
        row_count = 0
        row = start_row
        
        while True:
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                break
            if cell_value is not None:
                row_count += 1
            row += 4
        return row_count
    
    def format_time(self, time):
        return time.strftime("%H:%M") if time is not None else ""
    
    def _restaurant_data(self):
        sheet = self.contentbook['Restaurant']
        restaurants = {}
        if sheet:
            # Count even rows
            row_count = self.count_rows(sheet, "B", 15)

            # Loop through the rows and enter data
            cell_start = 15
            
            for i in range(row_count):
                # Get data from excel file
                # rank = i + 1
                rank = i + 1
                
                # Restuarant name
                rt_name = sheet[f"B{cell_start}"].value
                
                # Opening Hours information
                morning_open = sheet[f"L{cell_start}"].value
                morning_close = sheet[f"M{cell_start}"].value
                evening_open = sheet[f"N{cell_start + 1}"].value
                evening_close = sheet[f"O{cell_start + 1}"].value

                # format time object with function
                morning_open_formatted = self.format_time(morning_open)
                morning_close_formatted = self.format_time(morning_close)
                evening_open_formatted = self.format_time(evening_open)
                evening_close_formatted = self.format_time(evening_close)

                # condition for write open time
                if not morning_open and not morning_close and not evening_open and not evening_close:
                    open_hour = ""
                elif not evening_open and not evening_close:
                    open_hour = f'{morning_open_formatted}-{morning_close_formatted}'
                elif not morning_open and not morning_close:
                    open_hour = f'{evening_open_formatted}-{evening_close_formatted}'
                elif not morning_close and not evening_open:
                    open_hour = f'{morning_open_formatted}-{evening_close_formatted}'
                else:
                    open_hour = f'{morning_open_formatted}-{morning_close_formatted}/{evening_open_formatted}-{evening_close_formatted}'

                # Cooking Type Dropdown
                cook_type = sheet[f"P{cell_start}"].value
                
                # Open information Tickbox
                mid_mon = sheet[f"E{cell_start}"].value
                mid_tue = sheet[f"F{cell_start}"].value
                mid_wed = sheet[f"G{cell_start}"].value
                mid_thu = sheet[f"H{cell_start}"].value 
                mid_fri = sheet[f"I{cell_start}"].value
                mid_sat = sheet[f"J{cell_start}"].value
                mid_sun = sheet[f"K{cell_start}"].value
                eve_mon = sheet[f"E{cell_start + 1}"].value
                eve_tue = sheet[f"F{cell_start + 1}"].value
                eve_wed = sheet[f"G{cell_start + 1}"].value
                eve_thu = sheet[f"H{cell_start + 1}"].value 
                eve_fri = sheet[f"I{cell_start + 1}"].value
                eve_sat = sheet[f"J{cell_start + 1}"].value
                eve_sun = sheet[f"K{cell_start + 1}"].value

                # Payment Option Tickbox
                cash =          sheet[f"R{cell_start}"].value
                credit_card =   sheet[f"S{cell_start}"].value
                check =         sheet[f"T{cell_start}"].value
                other =         sheet[f"U{cell_start}"].value

                # Chef -> skip
                
                # price -> skip
                
                # Max seats
                max_seats = sheet[f"Q{cell_start}"].value
                if max_seats == None:
                    max_seats = 0
                    
                # Service option
                full_board =    sheet[f"V{cell_start}"].value
                half_board =    sheet[f"W{cell_start}"].value
                wheel_chair =   sheet[f"X{cell_start}"].value
                air_con =       sheet[f"Y{cell_start}"].value
                smoking =       sheet[f"Z{cell_start}"].value
                view =          sheet[f"AA{cell_start}"].value
                thematic =      sheet[f"AB{cell_start}"].value
                meal_pool =     sheet[f"AC{cell_start}"].value
                pet_allow =     sheet[f"AD{cell_start}"].value
                terrace =       sheet[f"AE{cell_start}"].value
                
                # Classifications & labels -> tickbox
                michelin_1 =    sheet[f"AF{cell_start}"].value
                michelin_bib =  sheet[f"AI{cell_start}"].value
                michelin_2 =    sheet[f"AG{cell_start}"].value
                aaa_guide =     sheet[f"AJ{cell_start}"].value
                michelin_3 =    sheet[f"AH{cell_start}"].value
                
                # Menus -> tickbox
                children =      sheet[f"AK{cell_start}"].value
                salt_free =     sheet[f"AL{cell_start}"].value
                delight =       sheet[f"AM{cell_start}"].value
                vegetarian =    sheet[f"AN{cell_start}"].value
                halal =         sheet[f"AQ{cell_start}"].value
                brunch =        sheet[f"AO{cell_start}"].value
                gluten_free =   sheet[f"AP{cell_start}"].value
                kosher =        sheet[f"AR{cell_start}"].value
                
                # Desctiption
                description = sheet[f"D{cell_start + 2}"].value
                
                # Ongoing Service -> skip
                
                # Parameter for central use -> skip
        
                # add to dict
                restaurants[rt_name] = [{'open_hour': open_hour},
                                        {'rank': rank},
                                        {'cook_type': cook_type},
                                        {'Open information': [
                                            mid_mon,
                                            mid_tue,
                                            mid_wed,
                                            mid_thu, 
                                            mid_fri, 
                                            mid_sat, 
                                            mid_sun, 
                                            eve_mon, 
                                            eve_tue, 
                                            eve_wed, 
                                            eve_thu, 
                                            eve_fri, 
                                            eve_sat, 
                                            eve_sun 
                                            ]},
                                        {'Payment Option': [
                                            cash,
                                            credit_card,
                                            check,
                                            other       
                                            ]},
                                        {'Max seats': max_seats},
                                        {'Service option': [
                                            full_board,
                                            half_board, 
                                            wheel_chair,
                                            air_con,
                                            smoking ,
                                            view,
                                            thematic,
                                            meal_pool,
                                            pet_allow,
                                            terrace
                                            ]},
                                        {'Classifications': [
                                            michelin_1,
                                            michelin_bib,
                                            michelin_2,
                                            aaa_guide,
                                            michelin_3 
                                            ]},
                                        {'Menus': [
                                            children,
                                            salt_free,
                                            delight,
                                            vegetarian,
                                            halal,
                                            brunch,
                                            gluten_free,
                                            kosher
                                        ]},
                                        {'Description': description}
                                        ]

                cell_start += 4
        self.restaurants = restaurants
    
    def _resaurant_description(self):
        temp = {}
        for key, value in self.restaurants.items():
            temp[key] = value[9]['Description']
    
        self.resaurant_description = pd.DataFrame(list(temp.items()), columns=['rt_name', 'description'])
        
    def _bar_data(self):
        sheet_name = "Bar"  
        bars = {}
        try:
            # Load Excel file and select the sheet
            sheet = self.contentbook[sheet_name]
            
            if sheet:
                # Count item(s) in the content book
                result = self.bar_count_rows(sheet, "D", 14)
                row_count, cell_record = result
                
                # Loop through the rows and enter data
                for i, cell in enumerate(cell_record):
                    # Get data from excel file
                    
                    bar_code = str(sheet[f"D{cell}"].value).split()
                    bar_code = str(bar_code[0])
                    
                    # bar name
                    bar_name = sheet[f"B{cell}"].value
                    
                    # Opening Hours information
                    open = sheet[f"L{cell}"].value
                    close = sheet[f"M{cell}"].value
                    
                    # format time object with function
                    open_formatted = self.format_time(open)
                    close_formatted = self.format_time(close)

                    # for write open time
                    open_hour = f'{open_formatted}-{close_formatted}'
                    
                    # Max seats
                    max_seats = sheet[f"N{cell}"].value
                    if max_seats == None:
                        max_seats = 0
                    
                    # Average Price -> skip
                    
                    # Service Tickbox
                    pet_allow =     sheet[f"O{cell}"].value
                    room_service =  sheet[f"P{cell}"].value
                    light_meal =    sheet[f"Q{cell}"].value
                    music =         sheet[f"R{cell}"].value
                    happy_hour =    sheet[f"S{cell}"].value
                    
                    # Opening Date -> skip 5 tabs
                    
                    # Open information
                    mon = sheet[f"E{cell}"].value
                    tue = sheet[f"F{cell}"].value
                    wed = sheet[f"G{cell}"].value
                    thu =  sheet[f"H{cell}"].value
                    fri =  sheet[f"I{cell}"].value
                    sat =  sheet[f"J{cell}"].value
                    sun =  sheet[f"K{cell}"].value
                    
                    # rank = i
                    rank = i + 1
                    
                    description = sheet[f"E{cell + 1}"].value
                    if description == None:
                        description = sheet[f"E{cell + 2}"].value
                    
                    # add to dict
                    bars[bar_name] = [{'Code': bar_code},
                                            {'Name': bar_name},
                                            {'Opening hours': open_hour},
                                            {'Max seats': max_seats},
                                            {'Services': [
                                                pet_allow,
                                                room_service,
                                                light_meal,
                                                music,
                                                happy_hour   
                                            ]},
                                            {'Open Information': [
                                                mon,
                                                tue,
                                                wed,
                                                thu,
                                                fri,
                                                sat,
                                                sun
                                            ]},
                                            {'Rank': rank},
                                            {'Description': description} 
                    ]
                    
        except Exception as e:
            print(f"An error occurred: {str(e)}")
        
        finally:
            self.bars = bars
    
    def _meeting_room_data(self):
        sheet_name = "Meeting Room"  
        meeting_room = {}
        
        try:
            sheet = self.contentbook[sheet_name]
            if sheet:
                # Count even rows
                even_row_count = self.count_even_rows(sheet, "C", 12)

                # Loop through the rows and enter data
                cell_start = 12
                for _ in range(even_row_count):
                    data = [
                            str(sheet[f"C{cell_start}"].value).strip(),
                            sheet[f"G{cell_start}"].value,
                            sheet[f"F{cell_start}"].value,
                            sheet[f"H{cell_start}"].value,
                            sheet[f"I{cell_start}"].value,
                            sheet[f"L{cell_start}"].value,
                            sheet[f"N{cell_start}"].value,
                            sheet[f"J{cell_start}"].value,
                            sheet[f"M{cell_start}"].value,
                            sheet[f"K{cell_start}"].value,
                            sheet[f"O{cell_start}"].value,
                            sheet[f"Q{cell_start}"].value,
                            sheet[f"P{cell_start}"].value,
                            sheet[f"R{cell_start}"].value
                            ]
                    tickbox = [sheet[f"S{cell_start}"].value,
                            sheet[f"T{cell_start}"].value,
                            sheet[f"U{cell_start}"].value,
                            sheet[f"V{cell_start}"].value]
                    description = sheet[f"E{cell_start + 1}"].value
                    cell_start += 2
                    
                    # replace None with 0
                    data = [self.replace_none_with_zero(value) for value in data]
                    data[1] = int(round(data[1], 0))
                    data[3] = round(data[3], 2)
                    name = data[0]
                    
                    meeting_room[name] = [data, tickbox, description]
                    
        except Exception as e:
            print(f"An error occurred: {str(e)}")
            
        finally:
            self.meeting_room = meeting_room
           
    # Function to count even rows until an empty cell is encountered                
    def count_even_rows(sef, sheet, column_letter, start_row):
        even_row_count = 0
        row = start_row
        while True:
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                break
            if row % 2 == 0:
                even_row_count += 1
            row += 2
        
        return even_row_count
    
    def _mean_of_access_data(self):
        sheet_name = "Address&Setup"
        sheet = self.contentbook[sheet_name]
        access_data = {}
        if sheet:
            self.hotel_direction = sheet['C84'].value 
            for i in range(4):
                if sheet[f'E{149 + i}'].value != None:
                    pattern = r'([A-Z]+) -'
                    code = re.findall(pattern, sheet[f'C{149 + i}'].value)[0]
                    data = [sheet[f'E{149 + i}'].value,
                            sheet[f'H{149 + i}'].value,
                            sheet[f'I{149 + i}'].value,
                            sheet[f'K{149 + i}'].value]
                    access_data[code] = data
                    
        self.mean_of_access = access_data
    
    def _main_attractions_data(self):
        sheet_name = "Main Attractions"
        sheet = self.contentbook[sheet_name]
        data = {}
        if sheet:
            # Loop over data file
            cell_start = 11
            # Check Unit Select
            if sheet['E8'].value == 'Km':
                self.unit_select = 'hotelIp.kilometerDistance'
            else:
                self.unit_select = 'hotelIp.milesDistance'
            
            for i in range(18):
                if sheet[f'C{cell_start + i}'].value != None:
                    # regex for extract text
                    pattern = r'([^:]+):'
                    
                    # data from excel
                    code = re.findall(pattern, sheet[f'B{cell_start + i}'].value)[0]
                    if sheet[f'G{cell_start + i}'].value != None:
                        ofi = self.extract_capital_letters(sheet[f'G{cell_start + i}'].value)
                    else:
                        ofi = None
                    shuttle = str(sheet[f'D{cell_start + i}'].value)
                    shuttle_service_type = str(sheet[f'E{cell_start + i}'].value)
                    distance = (sheet[f'I{cell_start + i}'].value)
                    minute_walk = (sheet[f'J{cell_start + i}'].value)
                    minute_drive = (sheet[f'K{cell_start + i}'].value)
                    name = (sheet[f'C{cell_start + i}'].value)
                    
                    data[code] = [name, ofi, shuttle, shuttle_service_type, distance, minute_walk, minute_drive]
        
        self.main_attractions = data
    
    def _surrounding_attracition_data(self):
        cell_list = []
        for i in range(11, 29):
            cell_list.append(f'C{i}')

        # Read the Excel file, skipping the first 12 rows, and using the 13th row as column headers
        df = pd.read_excel(self.filepath, 
                        sheet_name='Other Attractions', 
                        header=12, 
                        )
        
        df.drop(df.columns[0:2], axis=1, inplace=True)
        df.drop(df.columns[7], axis=1, inplace=True)

        # Rename column
        columns_name = ['name', 'shuttle', 'shuttle_service', 'orientation', 'distance', 'time_walk', 'time_drive']
        df.columns = columns_name
          
        # df for COMP
        df_comp = df.iloc[0:9]
        self.comp = df_comp.dropna(subset=[df_comp.columns[0]])

        # df for CONG
        df_cong = df.iloc[10:18]
        self.cong = df_cong.dropna(subset=[df_cong.columns[0]])

        # df for EXHI
        df_exhi = df.iloc[19:27]
        self.exhi = df_exhi.dropna(subset=[df_exhi.columns[0]])
    
        # df for EXPO
        df_expo = df.iloc[28:34]
        self.expo = df_expo.dropna(subset=[df_expo.columns[0]])

    # Name ACCOR Standard check
    def accor_name(self, title, input_name, input_surname) -> str:
        """
        Format Accordant Name

        This function takes a title, a name, and a surname as input and returns a formatted
        accordant name string with the first letter of the name capitalized and the surname in uppercase.

        Parameters:
        - title: str
            The title to be included in the formatted name.
        - input_name: str
            The input name to be formatted, with the first letter capitalized and the rest in lowercase.
        - input_surname: str
            The input surname to be formatted, converted to uppercase.

        Returns:
        str
            A formatted accordant name string in the format '{title} {formatted_name} {formatted_surname}'.

        Example:
        >>> accor_name("Mr.", "john", "doe")
        'Mr. John DOE'
        >>> accor_name("Ms.", "jane", "smith")
        'Ms. Jane SMITH'
        >>> accor_name("Dr.", "robert", "johnson")
        'Dr. Robert JOHNSON'
        """
        name = input_name[0].upper() + input_name[1:].lower()
        surname = input_surname.upper()
        return f'{title} {name} {surname}'
    
    # replace None value in list with zero
    def replace_none_with_zero(self, value):
        """
        Replace None with Zero

        This function takes a value as input and returns 0 if the value is None,
        otherwise, it returns the original value.

        Parameters:
        - value: Any | None
            The input value that may be None.

        Returns:
        int or Any
            If the input value is None, 0 is returned. Otherwise, the original value is returned.

        Example:
        >>> replace_none_with_zero(None)
        0
        >>> replace_none_with_zero(42)
        42
        >>> replace_none_with_zero("Hello")
        'Hello'
        """
        return 0 if value is None else value
    
    def bar_count_rows(self, sheet, column_letter, start_row):
        # Function to count even rows until an empty cell is encountered
        # This function use for search and confirm item(s) in the content book
        row_count = 0
        row = start_row
        search = 0
        cell_record = []
        while search <= 5:
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                search += 1
            if cell_value is not None:
                row_count += 1
                search = 0
                cell_record.append(row)
            row += 1
        return row_count, cell_record

    def _product_lib_df(self):
        # Load product library
        csv_path = 'products_lib.csv'
        self.product_lib_df = pd.read_csv(
            csv_path,
            header=0,
            sep=';'
        )
   
class ContentBook_v10(ContentBook):
    def __init__(self, filepath):
        super().__init__(filepath)
        
    def _address_setup_data(self):
        # Local Star
        sheet_name = 'Address&Setup'
        sheet = self.contentbook[sheet_name]
        
        # data for Address and Setup page
        self.hotel_rid = sheet['K4'].value
        self.hotel_name = sheet['C4'].value
        self.brand = sheet['C6'].value
        self.chain = sheet['J6'].value
        self.hotel_commercial_name = sheet['D32'].value
        self.open_date = sheet['K10'].value
        self.address1 = sheet['C34'].value
        self.address2 = sheet['J34'].value
        self.address3 = sheet['C37'].value
        self.state = sheet['J39'].value
        self.resa_phone = sheet['K43'].value
        self.resa_fax = sheet['K45'].value
        self.logging_type = sheet['C47'].value
        self.zip_code = sheet['K37'].value
        self.phone_country_code = sheet['C41'].value
        self.hotel_email = sheet['C45'].value
        self.code_place = sheet['C49'].value
        self.city = sheet['C39'].value
        self.phone = sheet['D41'].value
        self.fax = sheet['D43'].value
        self.country = sheet['J41'].value
    
        # List data for general page
        self.construction_date = sheet['C57'].value
        self.reno_date = sheet['J57'].value
        self.distribute_tars_date = sheet['D12'].value
        self.lodging_type = sheet['J63'].value
        self.environment = sheet['C61'].value
        self.location = sheet['J59'].value
        self.currency = sheet['C59'].value
        self.nb_lifts = sheet['K65'].value
        self.nb_rooms = sheet['C63'].value
        self.nb_floors = sheet['C65'].value
        
        # rating data
        self.local_rating = sheet['G65'].value
        self.north_star = None   

        # GM name
        gm_title = sheet['E53'].value
        gm_name = sheet['J55'].value
        gm_surname = sheet['J53'].value

        # GM name
        self.gm = self.accor_name(gm_title, gm_name, gm_surname)
        
        # GPS not send
        self.gps = sheet['K24'].value
        
        # Hotel Currency 3 digits
        self.currency_code = self.extract_currency(self.currency)
        
    def _mean_of_access_data(self):
        sheet_name = "Address&Setup"
        sheet = self.contentbook[sheet_name]
        access_data = {}
        if sheet:
            self.hotel_direction = sheet['C79'].value 
            for i in range(4):
                if sheet[f'E{149 + i}'].value != None:
                    pattern = r'([A-Z]+) -'
                    code = re.findall(pattern, sheet[f'C{144 + i}'].value)[0]
                    data = [sheet[f'E{144 + i}'].value,
                            sheet[f'H{144 + i}'].value,
                            sheet[f'I{144 + i}'].value,
                            sheet[f'K{144 + i}'].value]
                    access_data[code] = data
                    
        self.mean_of_access = access_data
        
class ContentBook_v16(ContentBook):
    def __init__(self, filepath):
        super().__init__(filepath)
    
    def _address_setup_data(self):
        # Local Star
        sheet_name = 'Address&Setup'
        sheet = self.contentbook[sheet_name]
        
        # data for Address and Setup page
        self.hotel_rid = sheet['K4'].value
        self.hotel_name = sheet['C4'].value
        self.brand = sheet['C6'].value
        self.chain = sheet['J6'].value
        self.hotel_commercial_name = sheet['D34'].value
        self.open_date = sheet['K10'].value
        self.address1 = sheet['C36'].value
        self.address2 = sheet['J36'].value
        self.address3 = sheet['C38'].value
        self.state = sheet['J41'].value
        self.resa_phone = sheet['K45'].value
        self.resa_fax = sheet['K47'].value
        self.logging_type = sheet['C49'].value
        self.zip_code = sheet['K39'].value
        self.phone_country_code = sheet['C43'].value
        self.hotel_email = sheet['C47'].value
        self.code_place = sheet['C51'].value
        self.city = sheet['C41'].value
        self.phone = sheet['D43'].value
        self.fax = sheet['D45'].value
        self.country = sheet['J43'].value
    
        # List data for general page
        self.construction_date = sheet['C59'].value
        self.reno_date = sheet['J59'].value
        self.distribute_tars_date = sheet['D12'].value
        self.lodging_type = sheet['J65'].value
        self.environment = sheet['C63'].value
        self.location = sheet['J61'].value
        self.currency = sheet['C61'].value
        self.nb_lifts = sheet['K67'].value
        self.nb_rooms = sheet['C65'].value
        self.nb_floors = sheet['C67'].value
        
        # rating data
        self.local_rating = sheet['C81'].value
        self.north_star = sheet['J81'].value
        
        # GM name
        gm_title = sheet['E55'].value
        gm_name = sheet['J57'].value
        gm_surname = sheet['J55'].value

        # GM name
        self.gm = self.accor_name(gm_title, gm_name, gm_surname)
        
        # GPS not send
        self.gps = sheet['K24'].value
        
        # Hotel Currency 3 digits
        self.currency_code = self.extract_currency(self.currency)
        
    def _mean_of_access_data(self):
        sheet_name = "Address&Setup"
        sheet = self.contentbook[sheet_name]
        access_data = {}
        if sheet:
            self.hotel_direction = sheet['C86'].value 
            for i in range(4):
                if sheet[f'E{151 + i}'].value != None:
                    pattern = r'([A-Z]+) -'
                    code = re.findall(pattern, sheet[f'C{151 + i}'].value)[0]
                    data = [sheet[f'E{151 + i}'].value,
                            sheet[f'H{151 + i}'].value,
                            sheet[f'I{151 + i}'].value,
                            sheet[f'K{151 + i}'].value]
                    access_data[code] = data
                    
        self.mean_of_access = access_data
        
        
class ContentBook_v17(ContentBook):
    def __init__(self, filepath):
        super().__init__(filepath)
        
    def _address_setup_data(self):
        # Local Star
        sheet_name = 'Address&Setup'
        sheet = self.contentbook[sheet_name]
        
        # data for Address and Setup page
        self.hotel_rid = sheet['K4'].value
        self.hotel_name = sheet['C4'].value
        self.brand = sheet['C6'].value
        self.chain = sheet['J6'].value
        self.hotel_commercial_name = sheet['D33'].value
        self.open_date = sheet['K10'].value
        self.address1 = sheet['C35'].value
        self.address2 = sheet['J35'].value
        self.address3 = sheet['C36'].value
        self.state = sheet['J40'].value
        self.resa_phone = sheet['K44'].value
        self.resa_fax = sheet['K46'].value
        self.logging_type = sheet['C48'].value
        self.zip_code = sheet['K38'].value
        self.phone_country_code = sheet['C42'].value
        self.hotel_email = sheet['C46'].value
        self.code_place = sheet['C50'].value
        self.city = sheet['C40'].value
        self.phone = sheet['D42'].value
        self.fax = sheet['D44'].value
        self.country = sheet['J42'].value
    
        # List data for general page
        self.construction_date = sheet['C58'].value
        self.reno_date = sheet['J58'].value
        self.distribute_tars_date = sheet['D12'].value
        self.lodging_type = sheet['J64'].value
        self.environment = sheet['C62'].value
        self.location = sheet['J60'].value
        self.currency = sheet['C60'].value
        self.nb_lifts = sheet['K66'].value
        self.nb_rooms = sheet['C64'].value
        self.nb_floors = sheet['C66'].value
        
        # rating data
        self.local_rating = sheet['C80'].value
        self.north_star = sheet['J80'].value
        
        # GM name
        gm_title = sheet['E54'].value
        gm_name = sheet['J56'].value
        gm_surname = sheet['J54'].value

        # GM name
        self.gm = self.accor_name(gm_title, gm_name, gm_surname)
        
        # GPS not send
        self.gps = sheet['K24'].value
        
        # Hotel Currency 3 digits
        self.currency_code = self.extract_currency(self.currency)
        
    def _mean_of_access_data(self):
        sheet_name = "Address&Setup"
        sheet = self.contentbook[sheet_name]
        access_data = {}
        if sheet:
            self.hotel_direction = sheet['C85'].value 
            for i in range(4):
                if sheet[f'E{150 + i}'].value != None:
                    pattern = r'([A-Z]+) -'
                    code = re.findall(pattern, sheet[f'C{150 + i}'].value)[0]
                    data = [sheet[f'E{150 + i}'].value,
                            sheet[f'H{150 + i}'].value,
                            sheet[f'I{150 + i}'].value,
                            sheet[f'K{150 + i}'].value]
                    access_data[code] = data
                    
        self.mean_of_access = access_data

