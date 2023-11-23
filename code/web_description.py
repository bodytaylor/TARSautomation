import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from web_driver_init import driver
from openpyxl import load_workbook
import time
from functions import *

# Tickbox in browser console
def tick_box(element):
    script = (f'var checkbox = document.getElementById("{element}"); checkbox.checked = !checkbox.checked;')
    return script
    
# Enter Translations
def product_description(code, type, description, marketing, hotel_rid):
    # Open Web Browser on translate page and wait for webpage load
    url = f'https://dataweb.accor.net/dotw-trans/translateHotelProduct!input.action?actionType=translate&hotelProduct.code={code}&hotelProduct.type.code={type}&hotelProduct.centralUse=true&'
    driver.get(url)
    page = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.TAG_NAME, 'h4'))
        )
    print(f'[INFO] - {page.text}')
    
    # Open description input box
    driver.execute_script(f"displayTranslateForm('translateInput','GB','{code}','{type}','{hotel_rid}','productsDescriptionsTable','true','true','GB','true')")
    # Wait for page to load
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, '//*[@id="translateHotelProductForm"]'))
        )
    
    # find description box
    description_box = driver.find_element(By.ID, "hotelProductTranslate.description")
    marketing_box =  driver.find_element(By.ID, "hotelProductTranslate.referenceLabel")
         
    # Get data from DataFrame and type in the discription box, do not change the secound locater!
    if description != None:
        description_box.clear()
        description_box.send_keys(description)
        time.sleep(0.5)

    # Get data from DataFrame and type in the marketing lable, box do not change the secound locater!
    if marketing != None:   
        marketing_box.clear()
        marketing_box.send_keys(marketing)
        time.sleep(0.5)

    # Click Translate!
    script = """
    document.getElementById('translateHotelProductForm.submitButton').click();
    """
    # response to popup        
    driver.execute_script(script)
    time.sleep(1)
    alert = driver.switch_to.alert
    alert.accept()
    
    # Wait for response
    WebDriverWait(driver, 7).until(
        EC.visibility_of_element_located((By.XPATH, '//*[@id="messages"]'))
        )
    
    try:
        action_message = driver.find_element(By.XPATH, '//*[@id="actionmessage"]')
        print(f'[INFO] - {action_message.text}')
    except:
        error_message = driver.find_element(By.XPATH, '//*[@id="errormessage"]')
        print(f'[INFO] - {error_message.text}')

# read description
def extract_data_from_excel(file_path):
    data_list = []

    # Load the Excel file
    wb = load_workbook(file_path, data_only=True)
    ws = wb['Main services']  # Specify the sheet name here

    # Initialize row and column indices
    current_row = 40
    current_col = 3  # Column C

    while current_row <= 125:
        l_value = ws.cell(row=current_row, column=12).value  # Column L

        if l_value == "Yes":
            code = str(ws.cell(row=current_row, column=current_col).value)[0:6].strip()
            description = None
            marketing = None

            # Check the next row for Description or Marketing
            for i in range(1, 3):
                cell_value = str(ws.cell(row=current_row + 1, column=current_col).value).split()
                
                if "Description." in cell_value:
                    description = ws.cell(row=current_row + 1, column=4).value
                elif "Marketing" in cell_value:
                    marketing = ws.cell(row=current_row + 1, column=4).value

            # Append the data to the list of dictionaries
            data_list.append({'Code': code, 'Description': description, 'Marketing': marketing})

        current_row += 1

    # Convert the list of dictionaries to a DataFrame
    df = pd.DataFrame(data_list)

    return df

# Search for product lib
def add_product(code, df):
    code = str(code).strip()
    if code in df['code'].values:
        result = df.loc[df['code'] == code].values[0].astype(str).tolist()
        element = f"addBasicElement('{result[0]}','{result[1]}','{result[2]}','{result[3]}','{result[4]}','{result[5]}');"
        return element
    else:
        return None

# find type of product
def find_type(df, code):
    code = str(code).strip()
    if code in df['code'].values:
        result = df.loc[df['code'] == code].values[0].astype(str).tolist()
        return result[2]
    else:
        return None

def add(hotel_rid):
    print("[WARNING] - Please Check Loading Form and Clean up all note from the code section!")
    print("[WARNING] - V.10 DWBUS -> DWBUS1")
    print("[WARNING] - V.10 DWCCL1 -> DWCCL")
    
    continue_program()
    # Load product library
    csv_path = 'products_lib.csv'
    product_lib_df = pd.read_csv(
        csv_path,
        header=0,
        sep=';'
    )

    file_path = f'hotel_workbook\{hotel_rid}\{hotel_rid}.xlsm'
    df = extract_data_from_excel(file_path)

    # open web
    from web_driver_init import driver
    driver.get('https://dataweb.accor.net/dotw-trans/productTabs!input.action')
    time.sleep(1)
    page = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="classicTabName"]'))
        )
    print(f'[INFO] - {page.text}')
    
    product_error = []
    product_not_found = []
    # Let's rolls!
    for index, row in df.iterrows():
        code = row['Code']
        product_to_add = add_product(code, df=product_lib_df)
        if product_to_add is None:
                product_not_found.append(code)
        else:
            driver.execute_script(product_to_add)
            # Wait for page to load
            WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="formTitle"]'))
            )
        
            # always yes on GDS
            tick_box_script = tick_box(element='hotelProduct.availableOnGDSMedia')
            driver.execute_script(tick_box_script)
                
            # Click add
            driver.execute_script('document.getElementById("hotelProduct.submitButton").click();')
            
            # Get response
            get_response(driver=driver, code=code, error=product_error)
        
    # Add description
    for index, row in df.iterrows():
        code = row['Code']
        if code not in product_not_found:
            type = find_type(code=code, df=product_lib_df)
            des = row['Description']
            mk_label = row['Marketing']
            product_description(code=code, description=des, marketing=mk_label, type=type, hotel_rid=hotel_rid)
            
    # Print error to user
    if len(product_error) != 0:
        for i in product_error:
            print(i)
    